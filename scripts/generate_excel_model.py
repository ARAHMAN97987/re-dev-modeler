#!/usr/bin/env python3
"""ZAN Financial Modeler — Full Excel Export with Live Formulas
8 sheets, all formula-driven. Blue=input, Black=formula, Green=cross-sheet."""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

def fill(a): return PatternFill(start_color=a, end_color=a, fill_type='solid')
FH_=fill('FF1F2937'); FS_=fill('FFF0FDF4'); FT_=fill('FFEEF2FF'); FW_=fill('FFFEFCE8')
FH =Font(name='Arial',bold=True, color='FFFFFFFF',size=10)
FHS=Font(name='Arial',bold=True, color='FFFFFFFF',size=9)
FTL=Font(name='Arial',bold=True, color='FFFFFFFF',size=12)
FSC=Font(name='Arial',bold=True, color='FF1F2937',size=10)
FB =Font(name='Arial',bold=True, color='FF000000',size=10)
FBS=Font(name='Arial',bold=True, color='FF000000',size=9)
FN =Font(name='Arial',bold=False,color='FF000000',size=10)
FNS=Font(name='Arial',bold=False,color='FF000000',size=9)
FI =Font(name='Arial',bold=False,color='FF2563EB',size=10)
FIS=Font(name='Arial',bold=False,color='FF2563EB',size=9)
FR =Font(name='Arial',bold=False,color='FF008000',size=9)
FRB=Font(name='Arial',bold=True, color='FF008000',size=9)
BB=Border(bottom=Side(style='medium',color='FF6B7280'))
TB=Border(top=Side(style='medium',color='FF6B7280'))
NUM='#,##0'; NUMN='#,##0;(#,##0);"-"'; PCT='0.0%'; DX='0.00"x"'
AC=Alignment(horizontal='center',vertical='center',wrap_text=True)

def sc(ws,r,c,v,f=None,fl=None,nf=None,al=None,bd=None):
    cell=ws.cell(row=r,column=c,value=v)
    if f:cell.font=f
    if fl:cell.fill=fl
    if nf:cell.number_format=nf
    if al:cell.alignment=al
    if bd:cell.border=bd
    return cell

def hdr(ws,r,c1,c2,t):
    for c in range(c1,c2+1): sc(ws,r,c,t if c==c1 else None,FTL,FH_)
    if c2>c1: ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)

def chdr(ws,r,hh,c1=1):
    for i,h in enumerate(hh): sc(ws,r,c1+i,h,FHS,FH_,al=AC)

def secr(ws,r,c1,c2,t):
    for c in range(c1,c2+1): sc(ws,r,c,t if c==c1 else None,FSC,FS_)

def generate(p, out):
    wb=Workbook(); wb.remove(wb.active)
    assets=p.get('assets',[]); phases=p.get('phases',[])
    pn=[ph['name'] for ph in phases]; na=len(assets); nph=len(pn)
    h=p.get('horizon',50); sy=p.get('startYear',2026); cur=p.get('currency','SAR')
    YC=lambda yi:5+yi; LC=YC(h-1)
    YR=lambda r:f'{CL(5)}{r}:{CL(LC)}{r}'

    # ═══════════ INPUTS ═══════════
    ws=wb.create_sheet("Inputs"); ws.sheet_properties.tabColor="3B82F6"
    ws.column_dimensions['A'].width=32; ws.column_dimensions['B'].width=18
    hdr(ws,1,1,3,f"{p.get('name','')} — Inputs / المدخلات")
    R=3
    def sec(t):
        nonlocal R; R+=1
        for c in range(1,4): sc(ws,R,c,t if c==1 else None,FSC,FS_)
        ws.merge_cells(start_row=R,start_column=1,end_row=R,end_column=3); R+=1
    def inp(l,v,nf=None,pct=False):
        nonlocal R; sc(ws,R,1,l,FN)
        sc(ws,R,2,v/100 if pct and isinstance(v,(int,float)) else v,FI,nf=nf or(PCT if pct else None))
        R+=1; return R-1

    sec("General / عام")
    rNM=inp("Project Name",p.get('name','')); rSY=inp("Start Year",sy); rHR=inp("Horizon",h)

    sec("Land / الأرض")
    rLT=inp("Land Type",p.get('landType','lease'))
    rLA=inp("Land Area (sqm)",p.get('landArea',0),NUM)
    rLR=inp("Annual Land Rent",p.get('landRentAnnual',0),NUM)
    rLE=inp("Rent Escalation %",p.get('landRentEscalation',0),PCT,True)
    rLN=inp("Escalation Every N Yrs",p.get('landRentEscalationEveryN',5))
    rLG=inp("Grace Period (yrs)",p.get('landRentGrace',0))
    rLTR=inp("Lease Term (yrs)",p.get('landRentTerm',50))
    rLP=inp("Purchase Price",p.get('landPurchasePrice',0),NUM)

    sec("CAPEX / التكاليف")
    rSF=inp("Soft Cost %",p.get('softCostPct',10),PCT,True)
    rCN=inp("Contingency %",p.get('contingencyPct',5),PCT,True)

    sec("Financing / التمويل")
    rLV=inp("Max LTV %",p.get('maxLtvPct',60),PCT,True)
    rFR=inp("Finance Rate %",p.get('financeRate',7),PCT,True)
    rTN=inp("Loan Tenor (yrs)",p.get('loanTenor',8))
    rDG=inp("Debt Grace (yrs)",p.get('debtGrace',3))
    rUF=inp("Upfront Fee %",p.get('upfrontFeePct',0.5),PCT,True)

    sec("Exit / التخارج")
    rES=inp("Exit Strategy",p.get('exitStrategy','sale'))
    rEY=inp("Exit Year (0=auto)",p.get('exitYear',0))
    rEM=inp("Exit Multiple (x rent)",p.get('exitMultiple',10))
    rEC=inp("Exit Cost %",p.get('exitCostPct',2),PCT,True)

    sec("Waterfall / الشلال")
    rGP=inp("GP Equity %",p.get('gpEquityPct',20),PCT,True)
    rPF=inp("Preferred Return %",p.get('prefReturnPct',15),PCT,True)
    rCR=inp("Carry %",p.get('carryPct',25),PCT,True)
    rLS=inp("LP Profit Split %",p.get('lpProfitSplitPct',70),PCT,True)

    sec("Fees / الرسوم")
    rSB=inp("Subscription Fee %",p.get('subscriptionFeePct',2),PCT,True)
    rMG=inp("Annual Mgmt Fee %",p.get('annualMgmtFeePct',0.9),PCT,True)
    rCU=inp("Annual Custody Fee (fixed)",p.get('annualCustodyFee',0),NUM)
    rDF=inp("Developer Fee % (CAPEX)",p.get('developerFeePct',10),PCT,True)
    rSTF=inp("Structuring Fee %",p.get('structuringFeePct',0),PCT,True)

    sec("Phase Allocation / التوزيع")
    rPA=R
    for i in range(nph):
        sc(ws,R,1,pn[i],FN); sc(ws,R,2,round(1.0/max(nph,1),4),FI,nf=PCT); R+=1

    # ═══════════ PROGRAM ═══════════
    ws2=wb.create_sheet("Program"); ws2.sheet_properties.tabColor="10B981"
    hdr(ws2,1,1,20,"Asset Program / برنامج المساحات")
    chdr(ws2,3,["Phase","Category","Name","Code","Notes","Plot Area","Footprint","GFA",
        "Rev Type","Eff%","Leasable","Rate/sqm","Op EBITDA","Esc%","Ramp","Occ%",
        "Cost/sqm","Start","Dur(mo)","Total CAPEX"])
    ww=[14,12,18,8,8,10,10,10,10,8,10,10,12,8,6,8,10,6,8,14]
    for i in range(20): ws2.column_dimensions[CL(i+1)].width=ww[i]

    PD=4 # program data start
    for i,a in enumerate(assets):
        r=PD+i
        sc(ws2,r,1,a.get('phase',''),FIS);sc(ws2,r,2,a.get('category',''),FIS)
        sc(ws2,r,3,a.get('name',''),FIS);sc(ws2,r,4,a.get('code',''),FIS);sc(ws2,r,5,'',FNS)
        sc(ws2,r,6,a.get('plotArea',0),FIS,nf=NUM);sc(ws2,r,7,a.get('footprint',0),FIS,nf=NUM)
        sc(ws2,r,8,a.get('gfa',0),FIS,nf=NUM);sc(ws2,r,9,a.get('revType','Lease'),FIS)
        sc(ws2,r,10,(a.get('efficiency',0)or 0)/100,FIS,nf=PCT)
        sc(ws2,r,11,f'=H{r}*J{r}',FNS,nf=NUM) # Leasable=GFA*Eff
        sc(ws2,r,12,a.get('leaseRate',0),FIS,nf=NUM)
        sc(ws2,r,13,a.get('opEbitda',0),FIS,nf=NUM)
        sc(ws2,r,14,(a.get('escalation',0)or 0)/100,FIS,nf=PCT)
        sc(ws2,r,15,a.get('rampUpYears',3),FIS)
        sc(ws2,r,16,(a.get('stabilizedOcc',100)or 100)/100,FIS,nf=PCT)
        sc(ws2,r,17,a.get('costPerSqm',0),FIS,nf=NUM)
        sc(ws2,r,18,a.get('constrStart',1),FIS);sc(ws2,r,19,a.get('constrDuration',12),FIS)
        sc(ws2,r,20,f"=H{r}*Q{r}*(1+Inputs!B{rSF})*(1+Inputs!B{rCN})",FBS,nf=NUM)
    PE=PD+na-1 # program data end

    # ═══════════ CALC ENGINE ═══════════
    ws0=wb.create_sheet("Calc"); ws0.sheet_properties.tabColor="059669"
    ws0.column_dimensions['A'].width=5;ws0.column_dimensions['B'].width=22
    ws0.column_dimensions['C'].width=12;ws0.column_dimensions['D'].width=14
    for yi in range(h): ws0.column_dimensions[CL(YC(yi))].width=12
    hdr(ws0,1,1,LC,f"{p.get('name','')} — Calculation Engine")
    sc(ws0,3,1,"#",FHS,FH_,al=AC);sc(ws0,3,2,"Asset",FHS,FH_,al=AC)
    sc(ws0,3,3,"Phase",FHS,FH_,al=AC);sc(ws0,3,4,"Total",FHS,FH_,al=AC)
    for yi in range(h): sc(ws0,3,YC(yi),sy+yi,FHS,FH_,nf='0',al=AC)

    # CAPEX per asset
    CR=5; sc(ws0,CR,2,"CAPEX SCHEDULE (Per Asset)",FB); CR+=1; CX=CR
    for i in range(na):
        r=CX+i;pr=PD+i
        sc(ws0,r,1,i+1,FNS);sc(ws0,r,2,f"=Program!C{pr}",FR);sc(ws0,r,3,f"=Program!A{pr}",FR)
        sc(ws0,r,4,f'=SUM({YR(r)})',FBS,nf=NUM)
        for yi in range(h):
            yr=yi+1
            sc(ws0,r,YC(yi),(f"=IF(OR(Program!H{pr}=0,Program!S{pr}=0,Program!T{pr}=0),0,"
                f"IF(AND({yr}>=Program!R{pr},{yr}<Program!R{pr}+CEILING(Program!S{pr}/12,1)),"
                f"Program!T{pr}/CEILING(Program!S{pr}/12,1),0))"),FNS,nf=NUM)
    CXE=CX+na-1; CXT=CXE+1
    sc(ws0,CXT,2,"TOTAL CAPEX",FB,FT_);sc(ws0,CXT,4,f'=SUM({YR(CXT)})',FB,FT_,nf=NUM)
    for yi in range(h):
        c=YC(yi);sc(ws0,CXT,c,f'=SUM({CL(c)}{CX}:{CL(c)}{CXE})',FBS,FT_,nf=NUM)

    # Revenue per asset
    sc(ws0,CXT+2,2,"INCOME SCHEDULE (Per Asset)",FB); RV=CXT+3
    for i in range(na):
        r=RV+i;pr=PD+i
        sc(ws0,r,1,i+1,FNS);sc(ws0,r,2,f"=Program!C{pr}",FR);sc(ws0,r,3,f"=Program!A{pr}",FR)
        sc(ws0,r,4,f'=SUM({YR(r)})',FBS,nf=NUM)
        for yi in range(h):
            yr=yi+1
            H=f"Program!H{pr}";R_=f"Program!R{pr}";S=f"Program!S{pr}"
            I=f"Program!I{pr}";K=f"Program!K{pr}";L=f"Program!L{pr}"
            M=f"Program!M{pr}";N=f"Program!N{pr}";O=f"Program!O{pr}";P=f"Program!P{pr}"
            OS=f"({R_}+CEILING({S}/12,1))"
            sc(ws0,r,YC(yi),(f"=IF(OR({H}=0,{R_}=0,{S}=0),0,"
                f"IF({yr}>={OS},"
                f"IF({I}=\"Operating\","
                f"{M}*(1+{N})^({yr}-{OS})*MIN(1,({yr}-{OS}+1)/MAX(1,{O})),"
                f"{K}*{L}*(1+{N})^({yr}-{OS})*MIN({P},{P}/MAX(1,{O})*({yr}-{OS}+1))),"
                f"0))"),FNS,nf=NUM)
    RVE=RV+na-1;RVT=RVE+1
    sc(ws0,RVT,2,"TOTAL INCOME",FB,FT_);sc(ws0,RVT,4,f'=SUM({YR(RVT)})',FB,FT_,nf=NUM)
    for yi in range(h):
        c=YC(yi);sc(ws0,RVT,c,f'=SUM({CL(c)}{RV}:{CL(c)}{RVE})',FBS,FT_,nf=NUM)

    # Land rent
    sc(ws0,RVT+2,2,"LAND RENT SCHEDULE",FB); LR=RVT+3
    sc(ws0,LR,2,"Total Land Rent",FBS);sc(ws0,LR,4,f'=SUM({YR(LR)})',FBS,nf=NUM)
    Br=f"Inputs!B{rLR}";Be=f"Inputs!B{rLE}";Bn=f"Inputs!B{rLN}"
    Bg=f"Inputs!B{rLG}";Bt=f"Inputs!B{rLTR}";Blt=f"Inputs!B{rLT}";Blp=f"Inputs!B{rLP}"
    for yi in range(h):
        yr=yi+1
        sc(ws0,LR,YC(yi),(f'=IF({Blt}="purchase",IF({yr}=1,{Blp},0),'
            f'IF({Blt}="partner",0,'
            f'IF({yr}<={Bg},0,IF({yr}>{Bt},0,'
            f'{Br}*(1+{Be})^INT(({yr}-{Bg}-1)/MAX(1,{Bn}))))))'),FNS,nf=NUM)

    LRP=LR+1 # per-phase land rent
    for pi in range(nph):
        r=LRP+pi;sc(ws0,r,2,f"  {pn[pi]}",FNS);sc(ws0,r,4,f'=SUM({YR(r)})',FNS,nf=NUM)
        for yi in range(h): sc(ws0,r,YC(yi),f"={CL(YC(yi))}{LR}*Inputs!B{rPA+pi}",FNS,nf=NUM)

    # Phase CAPEX/Revenue aggregation
    sc(ws0,LRP+nph+1,2,"CAPEX BY PHASE",FB); PC=LRP+nph+2
    for pi in range(nph):
        r=PC+pi;sc(ws0,r,2,pn[pi],FBS,FS_);sc(ws0,r,4,f'=SUM({YR(r)})',FBS,nf=NUM)
        for yi in range(h):
            c=YC(yi);sc(ws0,r,c,f'=SUMPRODUCT(($C${CX}:$C${CXE}="{pn[pi]}")*({CL(c)}{CX}:{CL(c)}{CXE}))',FNS,nf=NUM)

    sc(ws0,PC+nph+1,2,"INCOME BY PHASE",FB); PR=PC+nph+2
    for pi in range(nph):
        r=PR+pi;sc(ws0,r,2,pn[pi],FBS,FS_);sc(ws0,r,4,f'=SUM({YR(r)})',FBS,nf=NUM)
        for yi in range(h):
            c=YC(yi);sc(ws0,r,c,f'=SUMPRODUCT(($C${RV}:$C${RVE}="{pn[pi]}")*({CL(c)}{RV}:{CL(c)}{RVE}))',FNS,nf=NUM)

    # ═══════════ CASHFLOW ═══════════
    ws3=wb.create_sheet("CashFlow"); ws3.sheet_properties.tabColor="059669"
    ws3.column_dimensions['A'].width=30;ws3.column_dimensions['B'].width=8;ws3.column_dimensions['C'].width=15;ws3.column_dimensions['D'].width=14
    for yi in range(h): ws3.column_dimensions[CL(YC(yi))].width=12
    hdr(ws3,1,1,LC,"Unlevered Project Cash Flow / التدفقات النقدية")
    sc(ws3,3,1,"Line Item",FHS,FH_);sc(ws3,3,2,"Unit",FHS,FH_);sc(ws3,3,3,"Total",FHS,FH_);sc(ws3,3,4,"",FHS,FH_)
    for yi in range(h): sc(ws3,3,YC(yi),sy+yi,FHS,FH_,nf='0')

    cr=5
    for pi in range(nph):
        secr(ws3,cr,1,LC,pn[pi]);cr+=1
        ir=cr; sc(ws3,cr,1,"  Income",FNS);sc(ws3,cr,2,cur,FNS);sc(ws3,cr,3,f'=SUM({YR(cr)})',FNS,nf=NUM)
        for yi in range(h): sc(ws3,cr,YC(yi),f"=Calc!{CL(YC(yi))}{PR+pi}",FR,nf=NUM)
        cr+=1
        lr_=cr; sc(ws3,cr,1,"  Land Rent",FNS);sc(ws3,cr,2,cur,FNS);sc(ws3,cr,3,f'=SUM({YR(cr)})',FNS,nf=NUMN)
        for yi in range(h): sc(ws3,cr,YC(yi),f"=Calc!{CL(YC(yi))}{LRP+pi}*-1",FR,nf=NUMN)
        cr+=1
        cx_=cr; sc(ws3,cr,1,"  CAPEX",FNS);sc(ws3,cr,2,cur,FNS);sc(ws3,cr,3,f'=SUM({YR(cr)})',FNS,nf=NUMN)
        for yi in range(h): sc(ws3,cr,YC(yi),f"=Calc!{CL(YC(yi))}{PC+pi}*-1",FR,nf=NUMN)
        cr+=1
        nr=cr; sc(ws3,cr,1,f"  Net CF — {pn[pi]}",FBS);sc(ws3,cr,2,cur,FNS);sc(ws3,cr,3,f'=SUM({YR(cr)})',FBS,nf=NUMN)
        for yi in range(h):
            c=YC(yi);sc(ws3,cr,c,f"={CL(c)}{ir}+{CL(c)}{lr_}+{CL(c)}{cx_}",FBS,nf=NUMN,bd=BB)
        cr+=1
        sc(ws3,cr,1,f"  IRR — {pn[pi]}",FBS);sc(ws3,cr,2,"%",FNS)
        sc(ws3,cr,3,f'=IFERROR(IRR({CL(YC(0))}{nr}:{CL(LC)}{nr}),"-")',FBS,nf=PCT);cr+=2

    # Consolidated
    secr(ws3,cr,1,LC,"CONSOLIDATED / الموحّد");cr+=1
    ci=cr; sc(ws3,cr,1,"  Total Income",FNS);sc(ws3,cr,2,cur,FNS);sc(ws3,cr,3,f'=SUM({YR(cr)})',FNS,nf=NUM)
    for yi in range(h): sc(ws3,cr,YC(yi),f"=Calc!{CL(YC(yi))}{RVT}",FR,nf=NUM)
    cr+=1
    cl=cr; sc(ws3,cr,1,"  Total Land Rent",FNS);sc(ws3,cr,2,cur,FNS);sc(ws3,cr,3,f'=SUM({YR(cr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws3,cr,YC(yi),f"=Calc!{CL(YC(yi))}{LR}*-1",FR,nf=NUMN)
    cr+=1
    cc=cr; sc(ws3,cr,1,"  Total CAPEX",FNS);sc(ws3,cr,2,cur,FNS);sc(ws3,cr,3,f'=SUM({YR(cr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws3,cr,YC(yi),f"=Calc!{CL(YC(yi))}{CXT}*-1",FR,nf=NUMN)
    cr+=1
    cn=cr; sc(ws3,cr,1,"  Net CF — Consolidated",FB);sc(ws3,cr,2,cur,FNS);sc(ws3,cr,3,f'=SUM({YR(cr)})',FB,nf=NUMN)
    for yi in range(h):
        c=YC(yi);sc(ws3,cr,c,f"={CL(c)}{ci}+{CL(c)}{cl}+{CL(c)}{cc}",FBS,nf=NUMN,bd=BB)
    cr+=1
    cirr=cr; sc(ws3,cr,1,"  IRR",FB);sc(ws3,cr,2,"%",FNS)
    sc(ws3,cr,3,f'=IFERROR(IRR({CL(YC(0))}{cn}:{CL(LC)}{cn}),"-")',FB,nf=PCT);cr+=1
    cnpv=cr
    for d,l in [(0.10,"10%"),(0.12,"12%"),(0.14,"14%")]:
        sc(ws3,cr,1,f"  NPV @{l}",FNS);sc(ws3,cr,2,cur,FNS)
        sc(ws3,cr,3,f'=NPV({d},{CL(YC(0))}{cn}:{CL(LC)}{cn})',FNS,nf=NUM);cr+=1
    cum=cr; sc(ws3,cr,1,"  Cumulative CF",FNS);sc(ws3,cr,2,cur,FNS)
    for yi in range(h):
        c=YC(yi)
        sc(ws3,cr,c,f"={CL(c)}{cn}" if yi==0 else f"={CL(YC(yi-1))}{cum}+{CL(c)}{cn}",FNS,nf=NUMN)
    cr+=1
    sc(ws3,cr,1,"  Payback (years)",FBS)
    sc(ws3,cr,3,f'=IFERROR(MATCH(TRUE,INDEX({CL(YC(0))}{cum}:{CL(LC)}{cum}>0,0),0),"-")',FBS)

    # ═══════════ FUND MODEL ═══════════
    ws5=wb.create_sheet("Fund"); ws5.sheet_properties.tabColor="D97706"
    ws5.column_dimensions['A'].width=34;ws5.column_dimensions['B'].width=8;ws5.column_dimensions['C'].width=15;ws5.column_dimensions['D'].width=14
    for yi in range(h): ws5.column_dimensions[CL(YC(yi))].width=12
    hdr(ws5,1,1,LC,"Fund Model / نموذج الصندوق")
    sc(ws5,3,1,"Line Item",FHS,FH_);sc(ws5,3,2,"Unit",FHS,FH_);sc(ws5,3,3,"Total",FHS,FH_);sc(ws5,3,4,"",FHS,FH_)
    for yi in range(h): sc(ws5,3,YC(yi),sy+yi,FHS,FH_,nf='0')

    fr=5
    # Capital Structure
    secr(ws5,fr,1,LC,"CAPITAL STRUCTURE / هيكل رأس المال");fr+=1
    fDC=fr;sc(ws5,fr,1,"  Total Dev Cost",FNS);sc(ws5,fr,3,f"=Calc!D{CXT}",FR,nf=NUM);fr+=1
    fMD=fr;sc(ws5,fr,1,"  Max Debt (LTV)",FNS);sc(ws5,fr,3,f"=C{fDC}*Inputs!B{rLV}",FNS,nf=NUM);fr+=1
    fEQ=fr;sc(ws5,fr,1,"  Total Equity Required",FBS);sc(ws5,fr,3,f"=C{fDC}-C{fMD}",FBS,nf=NUM);fr+=1
    fGPE=fr;sc(ws5,fr,1,"  GP Equity",FNS);sc(ws5,fr,3,f"=C{fEQ}*Inputs!B{rGP}",FNS,nf=NUM);fr+=1
    fLPE=fr;sc(ws5,fr,1,"  LP Equity",FNS);sc(ws5,fr,3,f"=C{fEQ}*(1-Inputs!B{rGP})",FNS,nf=NUM);fr+=1
    fEXY=fr;sc(ws5,fr,1,"  Exit Year",FBS);sc(ws5,fr,3,f'=IF(Inputs!B{rEY}=0,Inputs!B{rTN}+1,Inputs!B{rEY})',FBS);fr+=2

    # Debt Schedule
    secr(ws5,fr,1,LC,"DEBT SCHEDULE / جدول الدين");fr+=1
    fDD=fr;sc(ws5,fr,1,"  Drawdown",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUM)
    for yi in range(h): sc(ws5,fr,YC(yi),f"=Calc!{CL(YC(yi))}{CXT}*Inputs!B{rLV}",FNS,nf=NUM)
    fr+=1

    # Committed debt = cumulative drawdown (needed for repayment)
    fCD=fr;sc(ws5,fr,1,"  Cumulative Drawdown",FNS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        sc(ws5,fr,YC(yi),f"=SUM({CL(YC(0))}{fDD}:{CL(YC(yi))}{fDD})" if yi>0 else f"={CL(YC(0))}{fDD}",FNS,nf=NUM)
    fr+=1

    fOB=fr;sc(ws5,fr,1,"  Opening Balance",FNS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        sc(ws5,fr,YC(yi),0 if yi==0 else f"={CL(YC(yi-1))}{fOB+2}",FNS,nf=NUM) # prev close
    fr+=1

    # Repayment: after grace, repay committed_debt / (tenor-grace) each year, until tenor
    fRP=fr;sc(ws5,fr,1,"  Repayment",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    grf=f"Inputs!B{rDG}";tnf=f"Inputs!B{rTN}"
    for yi in range(h):
        yr=yi+1;c=YC(yi)
        # At grace end, committed_debt is known. Repay = committed_at_grace / repay_years
        # committed at grace = cumulative drawdown at year=grace
        # Use: IF(yr>grace AND yr<=tenor, -CumDraw_at_grace / (tenor-grace), 0)
        # CumDraw at grace year → dynamic: we use the grace-year column of cumulative drawdown
        # Since grace is in an input cell, we need INDIRECT or a simpler approach:
        # Simpler: repay = -C(MaxDebt) / (tenor - grace) during repay period
        sc(ws5,fr,c,f"=IF(AND({yr}>{grf},{yr}<={tnf}),-C{fMD}/MAX(1,{tnf}-{grf}),0)",FNS,nf=NUMN)
    fr+=1

    fCB=fr;sc(ws5,fr,1,"  Closing Balance",FBS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"=MAX(0,{CL(c)}{fOB}+{CL(c)}{fDD}+{CL(c)}{fRP})",FBS,nf=NUM,bd=BB)
    fr+=1

    fIT=fr;sc(ws5,fr,1,"  Interest / Profit Charge",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"=-({CL(c)}{fOB}+{CL(c)}{fCB})/2*Inputs!B{rFR}",FNS,nf=NUMN)
    fr+=1

    fDS=fr;sc(ws5,fr,1,"  Total Debt Service",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUMN)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"={CL(c)}{fRP}+{CL(c)}{fIT}",FBS,nf=NUMN)
    fr+=1

    fDR=fr;sc(ws5,fr,1,"  DSCR",FBS);sc(ws5,fr,2,"x",FNS)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f'=IF(ABS({CL(c)}{fDS})=0,"-",CashFlow!{CL(c)}{ci}/ABS({CL(c)}{fDS}))',FBS,nf=DX)
    fr+=2

    # Fees
    secr(ws5,fr,1,LC,"FEES / الرسوم");fr+=1
    fSUBF=fr;sc(ws5,fr,1,"  Subscription Fee",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws5,fr,YC(yi),f"=-C{fEQ}*Inputs!B{rSB}" if yi==0 else 0,FNS,nf=NUMN)
    fr+=1
    fMGMF=fr;sc(ws5,fr,1,"  Management Fee",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws5,fr,YC(yi),f"=IF({yi+1}<=C{fEXY},-C{fEQ}*Inputs!B{rMG},0)",FNS,nf=NUMN)
    fr+=1
    fCUSTF=fr;sc(ws5,fr,1,"  Custody Fee",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws5,fr,YC(yi),f"=IF({yi+1}<=C{fEXY},-Inputs!B{rCU},0)",FNS,nf=NUMN)
    fr+=1
    fDEVF=fr;sc(ws5,fr,1,"  Developer Fee",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws5,fr,YC(yi),f"=IF(Calc!{CL(YC(yi))}{CXT}>0,-Calc!{CL(YC(yi))}{CXT}*Inputs!B{rDF},0)",FNS,nf=NUMN)
    fr+=1
    fSTRF=fr;sc(ws5,fr,1,"  Structuring Fee",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws5,fr,YC(yi),f"=-C{fDC}*Inputs!B{rSTF}" if yi==0 else 0,FNS,nf=NUMN)
    fr+=1
    fTFEE=fr;sc(ws5,fr,1,"  Total Fees",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUMN)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"={CL(c)}{fSUBF}+{CL(c)}{fMGMF}+{CL(c)}{fCUSTF}+{CL(c)}{fDEVF}+{CL(c)}{fSTRF}",FBS,nf=NUMN,bd=BB)
    fr+=2

    # Exit Proceeds
    secr(ws5,fr,1,LC,"EXIT PROCEEDS / عوائد التخارج");fr+=1
    fSTI=fr;sc(ws5,fr,1,"  Stabilized Annual Income",FNS);sc(ws5,fr,2,cur,FNS)
    sc(ws5,fr,3,f'=IFERROR(INDEX(CashFlow!{CL(YC(0))}{ci}:{CL(LC)}{ci},1,C{fEXY}),0)',FNS,nf=NUM)
    fr+=1
    fEXV=fr;sc(ws5,fr,1,"  Exit Value (Multiple × Income)",FBS);sc(ws5,fr,2,cur,FNS)
    sc(ws5,fr,3,f"=C{fSTI}*Inputs!B{rEM}",FBS,nf=NUM);fr+=1
    fEXC=fr;sc(ws5,fr,1,"  Exit Cost",FNS);sc(ws5,fr,2,cur,FNS)
    sc(ws5,fr,3,f"=C{fEXV}*Inputs!B{rEC}*-1",FNS,nf=NUMN);fr+=1
    fEXN=fr;sc(ws5,fr,1,"  Net Exit Proceeds",FBS);sc(ws5,fr,2,cur,FNS)
    sc(ws5,fr,3,f"=C{fEXV}+C{fEXC}",FBS,nf=NUM)
    # Place exit in the correct year column
    for yi in range(h):
        yr=yi+1;c=YC(yi)
        sc(ws5,fr,c,f"=IF({yr}=C{fEXY},C{fEXN},0)",FNS,nf=NUM)
    fr+=2

    # Levered Cash Flow
    secr(ws5,fr,1,LC,"LEVERED CASH FLOW / التدفقات المموّلة");fr+=1
    fUL=fr;sc(ws5,fr,1,"  Unlevered Net CF",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws5,fr,YC(yi),f"=CashFlow!{CL(YC(yi))}{cn}",FR,nf=NUMN)
    fr+=1
    fDL=fr;sc(ws5,fr,1,"  + Debt Drawdown",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUM)
    for yi in range(h): sc(ws5,fr,YC(yi),f"={CL(YC(yi))}{fDD}",FNS,nf=NUM)
    fr+=1
    fSL=fr;sc(ws5,fr,1,"  - Debt Service",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws5,fr,YC(yi),f"={CL(YC(yi))}{fDS}",FNS,nf=NUMN)
    fr+=1
    # Debt payoff at exit (remaining balance paid from exit proceeds)
    fDP=fr;sc(ws5,fr,1,"  - Debt Payoff at Exit",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h):
        yr=yi+1;c=YC(yi)
        sc(ws5,fr,c,f"=IF({yr}=C{fEXY},-{CL(c)}{fCB},0)",FNS,nf=NUMN)
    fr+=1
    fEP=fr;sc(ws5,fr,1,"  + Exit Proceeds",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUM)
    for yi in range(h): sc(ws5,fr,YC(yi),f"={CL(YC(yi))}{fEXN}",FNS,nf=NUM)
    fr+=1
    fFEL=fr;sc(ws5,fr,1,"  - Total Fees",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h): sc(ws5,fr,YC(yi),f"={CL(YC(yi))}{fTFEE}",FNS,nf=NUMN)
    fr+=1
    fLV=fr;sc(ws5,fr,1,"  Levered Net CF (to Equity)",FB);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FB,nf=NUMN)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"={CL(c)}{fUL}+{CL(c)}{fDL}+{CL(c)}{fSL}+{CL(c)}{fDP}+{CL(c)}{fEP}+{CL(c)}{fFEL}",FBS,nf=NUMN,bd=BB)
    fr+=1
    fLI=fr;sc(ws5,fr,1,"  Levered IRR",FB);sc(ws5,fr,2,"%",FNS)
    sc(ws5,fr,3,f'=IFERROR(IRR({CL(YC(0))}{fLV}:{CL(LC)}{fLV}),"-")',FB,nf=PCT);fr+=1
    for d,l in [(0.10,"10%"),(0.12,"12%")]:
        sc(ws5,fr,1,f"  Levered NPV @{l}",FNS);sc(ws5,fr,2,cur,FNS)
        sc(ws5,fr,3,f'=NPV({d},{CL(YC(0))}{fLV}:{CL(LC)}{fLV})',FNS,nf=NUM);fr+=1
    fr+=1

    # ═══════════ WATERFALL ═══════════
    secr(ws5,fr,1,LC,"WATERFALL DISTRIBUTION / توزيع الشلال");fr+=1
    # Cash Available for Distribution
    fCA=fr;sc(ws5,fr,1,"  Cash Available",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUM)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"=MAX(0,{CL(c)}{fLV})",FNS,nf=NUM)
    fr+=1

    # Equity Calls (negative = call from investors)
    fEC_=fr;sc(ws5,fr,1,"  Equity Calls (negative)",FNS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FNS,nf=NUMN)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"=MIN(0,{CL(c)}{fLV})",FNS,nf=NUMN)
    fr+=1

    # Unreturned Capital tracking
    fUC=fr;sc(ws5,fr,1,"  Unreturned Capital (opening)",FNS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        c=YC(yi)
        if yi==0: sc(ws5,fr,c,f"=ABS({CL(c)}{fEC_})",FNS,nf=NUM)
        else: sc(ws5,fr,c,f"=MAX(0,{CL(YC(yi-1))}{fUC}-{CL(YC(yi-1))}{fUC+1}+ABS({CL(c)}{fEC_}))",FNS,nf=NUM)
    fr+=1

    # T1: Return of Capital
    fT1=fr;sc(ws5,fr,1,"  T1: Return of Capital",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUM)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"=MIN({CL(c)}{fCA},{CL(c)}{fUC})",FBS,nf=NUM)
    fr+=1

    # Remaining after T1
    fR1=fr;sc(ws5,fr,1,"  Remaining after T1",FNS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"={CL(c)}{fCA}-{CL(c)}{fT1}",FNS,nf=NUM)
    fr+=1

    # Pref accrual (on unreturned capital)
    fPAC=fr;sc(ws5,fr,1,"  Pref Accrual (cumulative unpaid)",FNS);sc(ws5,fr,2,cur,FNS)
    prf=f"Inputs!B{rPF}"
    for yi in range(h):
        c=YC(yi)
        if yi==0:
            sc(ws5,fr,c,f"={CL(c)}{fUC}*{prf}",FNS,nf=NUM)
        else:
            # accrual = prev unpaid pref + this year's accrual on unreturned cap - what was paid last year
            sc(ws5,fr,c,f"=MAX(0,{CL(YC(yi-1))}{fPAC}-{CL(YC(yi-1))}{fPAC+1})+{CL(c)}{fUC}*{prf}",FNS,nf=NUM)
    fr+=1

    # T2: Preferred Return payment
    fT2=fr;sc(ws5,fr,1,"  T2: Preferred Return",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUM)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"=MIN({CL(c)}{fR1},{CL(c)}{fPAC})",FBS,nf=NUM)
    fr+=1

    # Remaining after T2
    fR2=fr;sc(ws5,fr,1,"  Remaining after T2",FNS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"={CL(c)}{fR1}-{CL(c)}{fT2}",FNS,nf=NUM)
    fr+=1

    # T3: GP Catch-up (carry% of cumulative pref paid goes to GP)
    fT3=fr;sc(ws5,fr,1,"  T3: GP Catch-Up",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUM)
    crf=f"Inputs!B{rCR}"
    for yi in range(h):
        c=YC(yi)
        # catch-up target = carry/(1-carry) * cumPref - cumCatchup_paid_so_far
        # simplified: MIN(remaining, remaining * carry/(1-carry)) if remaining > 0
        sc(ws5,fr,c,f"=MIN({CL(c)}{fR2},{CL(c)}{fR2}*{crf}/(1-{crf}))",FBS,nf=NUM)
    fr+=1

    # Remaining after T3
    fR3=fr;sc(ws5,fr,1,"  Remaining after T3",FNS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"={CL(c)}{fR2}-{CL(c)}{fT3}",FNS,nf=NUM)
    fr+=1

    # T4: Profit Split
    fT4LP=fr;sc(ws5,fr,1,"  T4: LP Profit Split",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUM)
    lsf=f"Inputs!B{rLS}"
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"={CL(c)}{fR3}*{lsf}",FBS,nf=NUM)
    fr+=1
    fT4GP=fr;sc(ws5,fr,1,"  T4: GP Profit Split",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUM)
    for yi in range(h):
        c=YC(yi);sc(ws5,fr,c,f"={CL(c)}{fR3}*(1-{lsf})",FBS,nf=NUM)
    fr+=2

    # LP/GP Total Distributions
    secr(ws5,fr,1,LC,"INVESTOR RETURNS / عوائد المستثمرين");fr+=1
    fLPD=fr;sc(ws5,fr,1,"  LP Total Distribution",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUM)
    for yi in range(h):
        c=YC(yi)
        # LP gets: ROC share (1-GP%) + Pref share (1-GP%) + T4 LP split
        lpsh=f"(1-Inputs!B{rGP})"
        sc(ws5,fr,c,f"={CL(c)}{fT1}*{lpsh}+{CL(c)}{fT2}*{lpsh}+{CL(c)}{fT4LP}",FBS,nf=NUM)
    fr+=1

    fLPCF=fr;sc(ws5,fr,1,"  LP Net Cash Flow",FB);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FB,nf=NUMN)
    for yi in range(h):
        c=YC(yi)
        sc(ws5,fr,c,f"={CL(c)}{fEC_}*(1-Inputs!B{rGP})+{CL(c)}{fLPD}",FBS,nf=NUMN,bd=BB)
    fr+=1

    fLPI=fr;sc(ws5,fr,1,"  LP IRR",FB);sc(ws5,fr,2,"%",FNS)
    sc(ws5,fr,3,f'=IFERROR(IRR({CL(YC(0))}{fLPCF}:{CL(LC)}{fLPCF}),"-")',FB,nf=PCT);fr+=1
    fLPM=fr;sc(ws5,fr,1,"  LP MOIC",FB);sc(ws5,fr,2,"x",FNS)
    sc(ws5,fr,3,f'=IFERROR(SUM({YR(fLPD)})/ABS(SUM({YR(fEC_)}))*(1-Inputs!B{rGP}),"-")',FB,nf=DX);fr+=1
    for d,l in [(0.10,"10%"),(0.12,"12%"),(0.14,"14%")]:
        sc(ws5,fr,1,f"  LP NPV @{l}",FNS);sc(ws5,fr,2,cur,FNS)
        sc(ws5,fr,3,f'=NPV({d},{CL(YC(0))}{fLPCF}:{CL(LC)}{fLPCF})',FNS,nf=NUM);fr+=1
    fLPCUM=fr;sc(ws5,fr,1,"  LP Cumulative CF",FNS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        c=YC(yi)
        sc(ws5,fr,c,f"={CL(c)}{fLPCF}" if yi==0 else f"={CL(YC(yi-1))}{fLPCUM}+{CL(c)}{fLPCF}",FNS,nf=NUMN)
    fr+=2

    fGPD=fr;sc(ws5,fr,1,"  GP Total Distribution",FBS);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FBS,nf=NUM)
    gpsh=f"Inputs!B{rGP}"
    for yi in range(h):
        c=YC(yi)
        sc(ws5,fr,c,f"={CL(c)}{fT1}*{gpsh}+{CL(c)}{fT2}*{gpsh}+{CL(c)}{fT3}+{CL(c)}{fT4GP}",FBS,nf=NUM)
    fr+=1

    fGPCF=fr;sc(ws5,fr,1,"  GP Net Cash Flow",FB);sc(ws5,fr,2,cur,FNS);sc(ws5,fr,3,f'=SUM({YR(fr)})',FB,nf=NUMN)
    for yi in range(h):
        c=YC(yi)
        sc(ws5,fr,c,f"={CL(c)}{fEC_}*{gpsh}+{CL(c)}{fGPD}",FBS,nf=NUMN,bd=BB)
    fr+=1
    fGPI=fr;sc(ws5,fr,1,"  GP IRR",FB);sc(ws5,fr,2,"%",FNS)
    sc(ws5,fr,3,f'=IFERROR(IRR({CL(YC(0))}{fGPCF}:{CL(LC)}{fGPCF}),"-")',FB,nf=PCT);fr+=1
    fGPM=fr;sc(ws5,fr,1,"  GP MOIC",FB);sc(ws5,fr,2,"x",FNS)
    sc(ws5,fr,3,f'=IFERROR(SUM({YR(fGPD)})/ABS(SUM({YR(fEC_)})*{gpsh}),"-")',FB,nf=DX);fr+=1
    for d,l in [(0.10,"10%"),(0.12,"12%"),(0.14,"14%")]:
        sc(ws5,fr,1,f"  GP NPV @{l}",FNS);sc(ws5,fr,2,cur,FNS)
        sc(ws5,fr,3,f'=NPV({d},{CL(YC(0))}{fGPCF}:{CL(LC)}{fGPCF})',FNS,nf=NUM);fr+=1
    fGPCUM=fr;sc(ws5,fr,1,"  GP Cumulative CF",FNS);sc(ws5,fr,2,cur,FNS)
    for yi in range(h):
        c=YC(yi)
        sc(ws5,fr,c,f"={CL(c)}{fGPCF}" if yi==0 else f"={CL(YC(yi-1))}{fGPCUM}+{CL(c)}{fGPCF}",FNS,nf=NUMN)

    # ═══════════ OUTPUTS ═══════════
    ws4=wb.create_sheet("Outputs"); ws4.sheet_properties.tabColor="8B5CF6"
    ws4.column_dimensions['A'].width=28;ws4.column_dimensions['B'].width=18;ws4.column_dimensions['C'].width=10
    hdr(ws4,1,1,3,"Project Outputs / المخرجات")
    chdr(ws4,3,["Metric","Value","Unit"])
    o=4
    def orow(lbl,fm,nf=NUM,u=cur):
        nonlocal o; sc(ws4,o,1,lbl,FNS);sc(ws4,o,2,fm,FRB if isinstance(fm,str) and fm.startswith("=") else FBS,nf=nf);sc(ws4,o,3,u,FNS);o+=1
    orow("Total CAPEX",f"=Calc!D{CXT}");orow(f"Total Income ({h}yr)",f"=Calc!D{RVT}")
    orow("Total Land Rent",f"=Calc!D{LR}");orow("Unlevered IRR",f"=CashFlow!C{cirr}",PCT,"%")
    orow("NPV @10%",f"=CashFlow!C{cnpv}");orow("Levered IRR",f"=Fund!C{fLI}",PCT,"%")
    orow("Max Debt",f"=Fund!C{fMD}");orow("Equity",f"=Fund!C{fEQ}")
    orow("Net Exit",f"=Fund!C{fEXN}");orow("Total Fees",f"=Fund!C{fTFEE}",NUMN)
    orow("LP IRR",f"=Fund!C{fLPI}",PCT,"%");orow("LP MOIC",f"=Fund!C{fLPM}",DX,"x")
    orow("GP IRR",f"=Fund!C{fGPI}",PCT,"%");orow("GP MOIC",f"=Fund!C{fGPM}",DX,"x")
    o+=1;sc(ws4,o,1,"Phase Breakdown",FB,FS_);sc(ws4,o,2,None,fl=FS_);sc(ws4,o,3,None,fl=FS_);o+=1
    chdr(ws4,o,["Phase","CAPEX","Income"]);o+=1
    for pi in range(nph):
        sc(ws4,o,1,pn[pi],FBS);sc(ws4,o,2,f"=Calc!D{PC+pi}",FR,nf=NUM);sc(ws4,o,3,f"=Calc!D{PR+pi}",FR,nf=NUM);o+=1

    # ═══════════ CHECKS ═══════════
    ws9=wb.create_sheet("Checks"); ws9.sheet_properties.tabColor="059669"
    ws9.column_dimensions['A'].width=45;ws9.column_dimensions['B'].width=15;ws9.column_dimensions['C'].width=10
    hdr(ws9,1,1,3,"Model Checks / الفحوصات")
    chdr(ws9,3,["Check","Diff","Status"])
    ck=4
    def chk(l,d):
        nonlocal ck;sc(ws9,ck,1,l,FNS);sc(ws9,ck,2,d,FNS,nf=NUM);sc(ws9,ck,3,f'=IF(ABS(B{ck})<1,"PASS","FAIL")',FBS);ck+=1
    px="+".join([f"Calc!D{PC+i}" for i in range(nph)])
    chk("CAPEX Total = Sum Phases",f"=Calc!D{CXT}-({px})")
    pv="+".join([f"Calc!D{PR+i}" for i in range(nph)])
    chk("Revenue Total = Sum Phases",f"=Calc!D{RVT}-({pv})")
    chk("D+E = Total Cost",f"=ABS(Fund!C{fMD}+Fund!C{fEQ}-Fund!C{fDC})")
    sc(ws9,ck,1,"No Negative GFA",FNS)
    sc(ws9,ck,2,f'=COUNTIF(Program!H{PD}:H{PE},"<0")',FNS,nf=NUM)
    sc(ws9,ck,3,f'=IF(B{ck}=0,"PASS","FAIL")',FBS);ck+=1
    sc(ws9,ck,1,"Active Assets Have Duration > 0",FNS)
    sc(ws9,ck,2,f"=SUMPRODUCT((Program!H{PD}:H{PE}>0)*(Program!S{PD}:S{PE}=0))",FNS,nf=NUM)
    sc(ws9,ck,3,f'=IF(B{ck}=0,"PASS","FAIL")',FBS);ck+=1
    sc(ws9,ck,1,"Phase Allocation = 100%",FNS)
    sc(ws9,ck,2,f"=ABS(SUM(Inputs!B{rPA}:B{rPA+nph-1})-1)",FNS,nf='0.0000')
    sc(ws9,ck,3,f'=IF(B{ck}<0.01,"PASS","FAIL")',FBS);ck+=1
    sc(ws9,ck,1,"Operating Assets Have EBITDA > 0",FNS)
    sc(ws9,ck,2,f'=SUMPRODUCT((Program!I{PD}:I{PE}="Operating")*(Program!M{PD}:M{PE}<=0))',FNS,nf=NUM)
    sc(ws9,ck,3,f'=IF(B{ck}=0,"PASS","WARN")',FBS);ck+=1
    chk("LP+GP Dist = Total Available",f"=ABS(SUM({YR(fLPD)})+SUM({YR(fGPD)})-SUM({YR(fCA)}))")
    sc(ws9,ck,1,"OVERALL STATUS",FB,FT_)
    sc(ws9,ck,2,"",None,FT_)
    sc(ws9,ck,3,f'=IF(COUNTIF(C4:C{ck-1},"FAIL")=0,"ALL PASS","ERRORS")',FB,FT_)

    # ═══════════ DASHBOARD ═══════════
    wdb=wb.create_sheet("Dashboard"); wdb.sheet_properties.tabColor="0F172A"
    wdb.column_dimensions['A'].width=24;wdb.column_dimensions['B'].width=20
    wdb.column_dimensions['C'].width=20;wdb.column_dimensions['D'].width=20
    wdb.column_dimensions['E'].width=16;wdb.column_dimensions['F'].width=14
    hdr(wdb,1,1,6,f"{p.get('name','')}");sc(wdb,2,1,"Project Financial Model Dashboard",FSC)
    dr=4;secr(wdb,dr,1,6,"KEY METRICS / المؤشرات الرئيسية");dr+=1
    chdr(wdb,dr,["Metric","Value","Unit","","",""]);dr+=1
    def dk(l,f,nf=NUM,u=cur):
        nonlocal dr;sc(wdb,dr,1,l,FNS);sc(wdb,dr,2,f,FRB,nf=nf);sc(wdb,dr,3,u,FNS);dr+=1
    dk("Total CAPEX",f"=Calc!D{CXT}");dk(f"Total Income ({h}yr)",f"=Calc!D{RVT}")
    dk("Total Land Rent",f"=Calc!D{LR}");dk("Unlevered IRR",f"=CashFlow!C{cirr}",PCT,"%")
    dk("NPV @10%",f"=CashFlow!C{cnpv}");dk("Levered IRR",f"=Fund!C{fLI}",PCT,"%")
    dk("Max Debt",f"=Fund!C{fMD}");dk("Total Equity",f"=Fund!C{fEQ}")
    dk("Net Exit",f"=Fund!C{fEXN}");dk("Total Fees",f"=Fund!C{fTFEE}",NUMN)
    dr+=1
    dk("LP IRR",f"=Fund!C{fLPI}",PCT,"%");dk("LP MOIC",f"=Fund!C{fLPM}",DX,"x")
    dk("GP IRR",f"=Fund!C{fGPI}",PCT,"%");dk("GP MOIC",f"=Fund!C{fGPM}",DX,"x")
    dr+=1
    secr(wdb,dr,1,6,"PHASE SUMMARY / ملخص المراحل");dr+=1
    chdr(wdb,dr,["Phase","CAPEX","Income","Assets","% of CAPEX",""]);dr+=1
    for pi in range(nph):
        sc(wdb,dr,1,pn[pi],FBS);sc(wdb,dr,2,f"=Calc!D{PC+pi}",FR,nf=NUM)
        sc(wdb,dr,3,f"=Calc!D{PR+pi}",FR,nf=NUM)
        sc(wdb,dr,4,f'=COUNTIF(Program!A{PD}:A{PE},"{pn[pi]}")',FNS)
        sc(wdb,dr,5,f"=IF(Calc!D{CXT}=0,0,Calc!D{PC+pi}/Calc!D{CXT})",FNS,nf=PCT);dr+=1
    sc(wdb,dr,1,"CONSOLIDATED",FB,FT_);sc(wdb,dr,2,f"=Calc!D{CXT}",FRB,FT_,nf=NUM)
    sc(wdb,dr,3,f"=Calc!D{RVT}",FRB,FT_,nf=NUM);sc(wdb,dr,4,na,FBS,FT_);sc(wdb,dr,5,1,FBS,FT_,nf=PCT)
    dr+=2
    secr(wdb,dr,1,6,"ASSET PROGRAM / برنامج الأصول");dr+=1
    chdr(wdb,dr,["No.","Asset","Category","GFA (sqm)","Cost/sqm","Total CAPEX"]);dr+=1
    for i in range(na):
        pr=PD+i;sc(wdb,dr,1,i+1,FNS);sc(wdb,dr,2,f"=Program!C{pr}",FR)
        sc(wdb,dr,3,f"=Program!B{pr}",FR);sc(wdb,dr,4,f"=Program!H{pr}",FR,nf=NUM)
        sc(wdb,dr,5,f"=Program!Q{pr}",FR,nf=NUM);sc(wdb,dr,6,f"=Program!T{pr}",FR,nf=NUM);dr+=1
    dr+=1
    secr(wdb,dr,1,6,"NPV ANALYSIS / تحليل القيمة الحالية");dr+=1
    chdr(wdb,dr,["Discount Rate","Project (Unlev)","Project (Lev)","LP","GP",""]);dr+=1
    for d,l in [(0.10,"10%"),(0.12,"12%"),(0.14,"14%")]:
        sc(wdb,dr,1,l,FBS)
        sc(wdb,dr,2,f"=NPV({d},CashFlow!{CL(YC(0))}{cn}:{CL(LC)}{cn})",FNS,nf=NUM)
        sc(wdb,dr,3,f"=NPV({d},Fund!{CL(YC(0))}{fLV}:{CL(LC)}{fLV})",FNS,nf=NUM)
        sc(wdb,dr,4,f"=NPV({d},Fund!{CL(YC(0))}{fLPCF}:{CL(LC)}{fLPCF})",FNS,nf=NUM)
        sc(wdb,dr,5,f"=NPV({d},Fund!{CL(YC(0))}{fGPCF}:{CL(LC)}{fGPCF})",FNS,nf=NUM);dr+=1

    # ═══════════ DOCS ═══════════
    wsd=wb.create_sheet("Docs"); wsd.sheet_properties.tabColor="6B7280"
    wsd.column_dimensions['A'].width=65
    hdr(wsd,1,1,1,"Documentation / التوثيق")
    for i,l in enumerate(["","Sheet Guide:","  Dashboard — KPIs + Phase Summary + Asset List + NPV Analysis",
        "  Inputs — Editable assumptions (blue text)","  Program — Asset table",
        "  Calc — Engine (all formulas)","  CashFlow — Unlevered cash flows by phase",
        "  Fund — Debt + Fees + Exit + Waterfall + LP/GP returns",
        "  Outputs — Summary KPIs","  Checks — Integrity checks (8 checks)","",
        "Color: Blue=Input, Black=Formula, Green=Cross-sheet",
        "","How to use:","  1. Edit Inputs sheet (blue cells)",
        "  2. Edit Program sheet (asset details)","  3. All other sheets recalculate automatically",
        "",f"Project: {p.get('name','')}",f"Horizon: {h}yr from {sy}",
        f"Assets: {na} | Phases: {', '.join(pn)}","Generated by ZAN Financial Modeler"]):
        sc(wsd,2+i,1,l,FN)

    # Freeze + order
    ws0.freeze_panes=f'{CL(YC(0))}4';ws3.freeze_panes=f'{CL(YC(0))}4';ws5.freeze_panes=f'{CL(YC(0))}4'
    desired=["Dashboard","Inputs","Program","Calc","CashFlow","Fund","Outputs","Checks","Docs"]
    for i,name in enumerate(desired):
        idx=wb.sheetnames.index(name);wb.move_sheet(name,offset=i-idx)
    wb.save(out); return out

if __name__=="__main__":
    with open(sys.argv[1]) as f: proj=json.load(f)
    print(f"Generated: {generate(proj, sys.argv[2])}")
